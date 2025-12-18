# -*- coding: utf-8 -*-
"""
Flask Application - Yeshiva Management System
מערכת Flask - מערכת ניהול ישיבה
"""

from flask import Flask, render_template, request, jsonify, send_file, make_response, session, redirect, url_for
from datetime import datetime, timedelta
from pyluach import dates
from services.database import YeshivaDatabase, ExamDatabase
from functools import wraps
import json
import os
import hashlib
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # Support Hebrew characters in JSON
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['TESSERACT_PATH'] = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'yeshiva-secret-key-2024-change-in-production')

# Default admin password (change this!)
ADMIN_PASSWORD_HASH = hashlib.sha256(os.environ.get('ADMIN_PASSWORD', 'yeshiva123').encode()).hexdigest()

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = YeshivaDatabase()
exam_db = ExamDatabase()

# ==================== AUTHENTICATION ====================

def login_required(f):
    """דקורטור לבדיקת התחברות"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    """דף התחברות"""
    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if password_hash == ADMIN_PASSWORD_HASH:
            session['logged_in'] = True
            session['login_time'] = datetime.now().isoformat()
            return redirect(url_for('dashboard'))
        else:
            error = 'סיסמה שגויה'
    
    return render_template('login.html', error=error, current_year=datetime.now().year)

@app.route('/logout')
def logout():
    """התנתקות"""
    session.clear()
    return redirect(url_for('login'))

# ==================== PAGES ====================

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    """לוח בקרה - Dashboard"""
    today = datetime.now().date()
    heb_date = dates.HebrewDate.from_pydate(today)

    students = db.get_all_students(include_inactive=False)
    total_students = len(students)

    # Get today's attendance
    attendance_data = db.get_attendance_for_date(today)
    present_count = len([a for a in attendance_data if a[2] == 'נוכח'])
    absent_count = len([a for a in attendance_data if a[2] == 'חסר'])

    if total_students > 0:
        attendance_percent = int((present_count / total_students) * 100)
    else:
        attendance_percent = 0

    stats = {
        'present': present_count,
        'absent': absent_count,
        'percent': attendance_percent,
        'total': total_students,
        'today': today.strftime('%d/%m/%Y'),
        'hebrew_date': heb_date.hebrew_date_string()
    }

    # נתונים דינמיים ללוח הבקרה
    weekly_attendance = db.get_weekly_attendance_by_day()
    low_attendance_students = db.get_low_attendance_students(threshold=80, days=30)
    upcoming_exams = exam_db.get_upcoming_exams(days=7)
    recent_activities = db.get_recent_activities(limit=10)

    return render_template('dashboard.html', 
                         stats=stats,
                         weekly_attendance=weekly_attendance,
                         low_attendance=low_attendance_students,
                         upcoming_exams=upcoming_exams,
                         activities=recent_activities)

@app.route('/students')
@login_required
def students():
    """ניהול תלמידים - Students Management"""
    return render_template('students.html')

@app.route('/attendance')
@login_required
def attendance():
    """סימון נוכחות - Attendance"""
    return render_template('attendance.html')

@app.route('/reports')
@login_required
def reports():
    """דוחות - Reports"""
    return render_template('reports.html')

@app.route('/demo-report')
def demo_report():
    """דמו עיצוב לדוח נוכחות"""
    return render_template('demo_report.html')

@app.route('/settings')
@login_required
def settings():
    """הגדרות - Settings"""
    return render_template('settings.html')

# ==================== API ENDPOINTS ====================

@app.route('/api/search')
@login_required
def api_global_search():
    """API: חיפוש גלובלי - תלמידים, מבחנים, דפים"""
    query = request.args.get('q', '').strip().lower()
    
    if len(query) < 1:
        return jsonify({'students': [], 'exams': []})
    
    # חיפוש תלמידים
    students = db.get_all_students(include_inactive=False)
    matching_students = []
    
    for student in students:
        student_id, first_name, last_name = student[0], student[1], student[2]
        current_grade = student[15] if len(student) > 15 else ''
        full_name = f"{first_name} {last_name}".lower()
        
        if query in full_name or query in first_name.lower() or query in last_name.lower():
            matching_students.append({
                'id': student_id,
                'first_name': first_name,
                'last_name': last_name,
                'grade': current_grade or ''
            })
    
    # הגבלה ל-10 תוצאות
    matching_students = matching_students[:10]
    
    return jsonify({
        'students': matching_students
    })

@app.route('/api/students')
def api_get_students():
    """API: קבלת רשימת תלמידים"""
    search_query = request.args.get('search', '').lower()
    grade_filter = request.args.get('grade', 'הכל')

    students = db.get_all_students(include_inactive=False)
    filtered = []

    for student in students:
        student_id, first_name, last_name, id_number, birth_date, address, city, \
            father_name, father_id_number, mother_name, mother_id_number, \
            father_phone, mother_phone, home_phone, entry_date, current_grade, \
            initial_grade, status, framework_type, notes, created_at, last_grade_update = student

        # Search filter
        full_name = f"{first_name} {last_name}".lower()
        if search_query and search_query not in full_name and search_query not in (id_number or ''):
            continue

        # Grade filter
        if grade_filter != 'הכל' and current_grade != grade_filter:
            continue

        filtered.append({
            'id': student_id,
            'first_name': first_name,
            'last_name': last_name,
            'name': f"{first_name} {last_name}",
            'grade': current_grade if current_grade else '-',
            'id_number': id_number or '',
            'birth_date_hebrew': birth_date or '',
            'address': address or '',
            'city': city or '',
            'father_name': father_name or '',
            'father_id_number': father_id_number or '',
            'mother_name': mother_name or '',
            'mother_id_number': mother_id_number or '',
            'home_phone': home_phone or '',
            'mobile_phone': father_phone or mother_phone or '',
            'status': status
        })

    return jsonify(filtered)

@app.route('/api/attendance/<date>')
@app.route('/api/attendance/<date>/<session>')
def api_get_attendance(date, session='שחרית'):
    """API: קבלת נוכחות לתאריך וסשן (תפילה או סדר לימוד)"""
    try:
        gregorian_date = datetime.strptime(date, '%Y-%m-%d').date()
        attendance_data = db.get_attendance_for_date(gregorian_date, session)

        heb_date = dates.HebrewDate.from_pydate(gregorian_date)

        students = db.get_all_students(include_inactive=False)
        total_students = len(students)
        present_count = len([a for a in attendance_data if a[2] == 'נוכח'])
        absent_count = len([a for a in attendance_data if a[2] == 'חסר'])
        late_count = len([a for a in attendance_data if a[2] == 'איחור'])
        unmarked = total_students - present_count - absent_count - late_count

        if total_students > 0:
            attendance_percent = int((present_count / total_students) * 100)
        else:
            attendance_percent = 0

        result = {
            'date': date,
            'hebrew_date': heb_date.hebrew_date_string(),
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'unmarked': unmarked,
            'percent': attendance_percent,
            'session': session,
            'students': []
        }

        # Add students with their attendance status
        for student in students:
            student_id = student[0]
            first_name = student[1]
            last_name = student[2]
            grade = student[15] if len(student) > 15 else 'א'  # current_grade is at index 15

            status = None
            late_time = None
            for att in attendance_data:
                if att[0] == student_id:
                    status = att[2]
                    late_time = att[3] if len(att) > 3 else None
                    break

            result['students'].append({
                'id': student_id,
                'name': f"{first_name} {last_name}",
                'status': status,
                'late_time': late_time,
                'grade': grade
            })

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/rapid-filling/<date>')
@app.route('/api/rapid-filling/<date>/<prayer>')
def api_get_rapid_filling(date, prayer='שחרית'):
    """API: קבלת רשימת תלמידים לסימון רץ - תומך בכל התפילות"""
    try:
        students = db.get_all_students(include_inactive=False)
        result = []
        for student in students:
            result.append({
                'id': student[0],
                'name': f"{student[1]} {student[2]}",
                'grade': student[15] if len(student) > 15 and student[15] else '-'
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/attendance/mark', methods=['POST'])
def api_mark_attendance():
    """API: סימון נוכחות - תומך בתפילות וסדרי לימוד"""
    try:
        data = request.json
        student_id = data.get('student_id')
        date = data.get('date')
        status_value = data.get('status')  # Can be: 1/0 (old format) or 'נוכח'/'חסר'/'איחור' (new format)
        session_type = data.get('session_type') or data.get('prayer_type', 'שחרית')  # תמיכה לאחור
        category = data.get('category', 'תפילה')

        gregorian_date = datetime.strptime(date, '%Y-%m-%d').date()

        # Convert to Hebrew date
        from services.date_service import HebrewDateConverter
        heb_date = HebrewDateConverter.get_hebrew_date(gregorian_date)

        # תמיכה לאחור: המרה מ-1/0 ל-'נוכח'/'חסר'
        if isinstance(status_value, int):
            status = 'נוכח' if status_value == 1 else 'חסר'
        else:
            status = status_value

        db.save_attendance(student_id, heb_date, date, status, session_type, category)

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/attendance/late-time', methods=['POST'])
def api_save_late_time():
    """API: שמירת שעת איחור"""
    try:
        data = request.json
        student_id = data.get('student_id')
        date = data.get('date')
        late_time = data.get('late_time')
        session_type = data.get('session_type', 'שחרית')
        category = data.get('category', 'תפילה')
        
        gregorian_date = datetime.strptime(date, '%Y-%m-%d').date()
        
        # Convert to Hebrew date
        from services.date_service import HebrewDateConverter
        heb_date = HebrewDateConverter.get_hebrew_date(gregorian_date)
        
        # שמירת שעת האיחור בבסיס הנתונים
        db.save_late_time(student_id, heb_date, date, late_time, session_type, category)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/classes')
def api_get_classes():
    """API: קבלת רשימת כיתות עם סטטיסטיקות"""
    students = db.get_all_students(include_inactive=False)
    class_counts = {}

    for student in students:
        current_grade = student[12]
        if current_grade not in class_counts:
            class_counts[current_grade] = 0
        class_counts[current_grade] += 1

    result = []
    for grade, count in sorted(class_counts.items()):
        result.append({
            'grade': grade if grade else 'לא מוגדר',
            'count': count
        })

    return jsonify(result)

@app.route('/api/sessions')
def api_get_all_sessions():
    """API: קבלת כל הסשנים (תפילות + סדרי לימוד)"""
    try:
        sessions = db.get_all_sessions()
        return jsonify(sessions)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/sessions/<date>')
def api_get_sessions_for_date(date):
    """API: קבלת סשנים פעילים לתאריך מסוים"""
    try:
        gregorian_date = datetime.strptime(date, '%Y-%m-%d').date()
        weekday = gregorian_date.weekday()  # 0=Monday, 6=Sunday

        # המרה ליום עברי: 0=ראשון, 6=שבת
        hebrew_weekday = (weekday + 1) % 7

        active_sessions = db.get_sessions_for_date(hebrew_weekday)

        return jsonify({
            'date': date,
            'weekday': hebrew_weekday,
            'sessions': active_sessions
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/sessions/category/<category>')
def api_get_sessions_by_category(category):
    """API: קבלת סשנים לפי קטגוריה (תפילה/לימוד)"""
    try:
        sessions = db.get_sessions_by_category(category)
        return jsonify(sessions)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/export/csv')
def api_export_csv():
    """API: יצוא לCSV"""
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        grade_filter = request.args.get('grade', 'הכל')

        start_date = datetime.strptime(start_date_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y').date()

        students = db.get_all_students(include_inactive=False)

        rows = [['שם מלא', 'שיעור', 'תעודת זהות', 'נוכחויות', 'העדרויות', 'אחוז']]

        for student in students:
            student_id, first_name, last_name, id_number, birth_date, address, city, \
                father_name, father_id_number, mother_name, mother_id_number, \
                father_phone, mother_phone, home_phone, entry_date, current_grade, \
                initial_grade, status, framework_type, notes, created_at, last_grade_update = student

            full_name = f"{first_name} {last_name}"
            grade = current_grade if current_grade else "-"

            if grade_filter != 'הכל' and current_grade != grade_filter:
                continue

            present = 0
            absent = 0
            late = 0

            current = start_date
            while current <= end_date:
                attendance_data = db.get_attendance_for_date(current, 'שחרית')
                for att in attendance_data:
                    if att[0] == student_id:
                        if att[2] == 'נוכח':
                            present += 1
                        elif att[2] == 'חסר':
                            absent += 1
                        elif att[2] == 'איחור':
                            late += 1
                        break
                current += timedelta(days=1)

            total_days = (end_date - start_date).days + 1
            percent = int((present / total_days * 100)) if total_days > 0 else 0

            rows.append([full_name, grade, id_number, present, absent, f"{percent}%"])

        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)

        return jsonify({
            'csv': output.getvalue(),
            'filename': f"דוח_נוכחות_{datetime.now().strftime('%Y%m%d')}.csv"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/reports/attendance')
def api_reports_attendance():
    """API: דוח נוכחות מסוכם לפי תלמיד בטווח תאריכים (JSON)"""
    try:
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        grade_filter = request.args.get('grade', 'הכל')

        if not start_date_str or not end_date_str:
            return jsonify({'error': 'חסר טווח תאריכים'}), 400

        start_date = datetime.strptime(start_date_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y').date()

        students = db.get_all_students(include_inactive=False)

        results = []
        for student in students:
            student_id, first_name, last_name, id_number, birth_date, address, city, \
                father_name, father_id_number, mother_name, mother_id_number, \
                father_phone, mother_phone, home_phone, entry_date, current_grade, \
                initial_grade, status, framework_type, notes, created_at, last_grade_update = student

            if grade_filter != 'הכל' and current_grade != grade_filter:
                continue

            present = 0
            absent = 0
            late = 0

            current = start_date
            while current <= end_date:
                attendance_data = db.get_attendance_for_date(current, 'שחרית')
                for att in attendance_data:
                    if att[0] == student_id:
                        if att[2] == 'נוכח':
                            present += 1
                        elif att[2] == 'חסר':
                            absent += 1
                        elif att[2] == 'איחור':
                            late += 1
                        break
                current += timedelta(days=1)

            total_days = (end_date - start_date).days + 1
            percent = int((present / total_days * 100)) if total_days > 0 else 0

            results.append({
                'id': student_id,
                'name': f"{first_name} {last_name}",
                'grade': current_grade if current_grade else '-',
                'id_number': id_number,
                'present': present,
                'absent': absent,
                'percent': percent
            })

        return jsonify({'rows': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/date/hebrew')
def api_date_hebrew():
    """API: המרת תאריך לועזי (DD/MM/YYYY) לעברי"""
    try:
        date_str = request.args.get('date')
        if not date_str:
            return jsonify({'error': 'חסר פרמטר date'}), 400
        pydate = datetime.strptime(date_str, '%d/%m/%Y').date()
        from services.date_service import HebrewDateConverter
        heb = HebrewDateConverter.get_hebrew_date(pydate)
        return jsonify({'hebrew': heb})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/date/gregorian')
def api_date_gregorian():
    """API: המרת תאריך עברי ללועזי (DD/MM/YYYY)"""
    try:
        hebrew_date_str = request.args.get('hebrew_date')
        if not hebrew_date_str:
            return jsonify({'error': 'חסר פרמטר hebrew_date'}), 400
        
        from pyluach import dates
        from services.date_service import HebrewDateConverter
        
        # ניסיון להמיר תאריך עברי ללועזי
        # נבדוק כמה פורמטים אפשריים
        try:
            # פורמט: "א׳ בשבט תשפ״ה"
            pydate = HebrewDateConverter.parse_hebrew_date(hebrew_date_str)
            if pydate:
                formatted = f"{pydate.day:02d}/{pydate.month:02d}/{pydate.year}"
                return jsonify({'gregorian': formatted})
        except:
            pass
            
        return jsonify({'error': 'לא ניתן להמיר תאריך זה'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/student-report')
@app.route('/student-report/<int:student_id>')
@login_required
def student_report_page(student_id=None):
    """עמוד הפקת דוח פרטי לתלמיד"""
    return render_template('student_report.html', student_id=student_id)

@app.route('/api/student-report-data/<int:student_id>')
def api_student_report_data(student_id):
    """API: קבלת כל נתוני התלמיד לדוח"""
    try:
        student = db.get_student_by_id(student_id)
        if not student:
            return jsonify({'error': 'תלמיד לא נמצא'}), 404
        
        # Get student exams - מחזיר dictionary לפי מקצוע
        exams_dict = db.get_student_exams(student_id)
        
        return jsonify({
            'id': student['id'],
            'name': f"{student['first_name']} {student['last_name']}",
            'grade': student.get('current_grade', ''),
            'birth_date': student.get('birth_date_hebrew', ''),
            'address': student.get('address', ''),
            'city': student.get('city', ''),
            'phone': student.get('home_phone', ''),
            'exams': exams_dict
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/student-report/preview', methods=['POST'])
def api_student_report_preview():
    """API: יצירת תצוגה מקדימה של הדוח"""
    try:
        data = request.json
        student_id = data.get('student_id')
        intro_text = data.get('intro_text', '')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        exams = data.get('exams', [])
        
        student = db.get_student_by_id(student_id)
        if not student:
            return jsonify({'error': 'תלמיד לא נמצא'}), 404
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y').date()
        
        # Get attendance summary and weekly data
        attendance_summary = db.get_student_attendance_summary(student_id, start_date, end_date)
        attendance_weekly = db.get_student_attendance_weekly(student_id, start_date, end_date)
        
        # Format exam data for horizontal table
        exam_students = [{
            'name': f"{student['first_name']} {student['last_name']}",
            'iyon': exams.get('iyon', ''),
            'bekiut': exams.get('bekiut', ''),
            'gemara_rashi': exams.get('gemara_rashi', ''),
            'chumash': exams.get('chumash', '')
        }]
        
        # Get Hebrew dates
        from services.date_service import HebrewDateConverter
        start_heb = HebrewDateConverter.get_hebrew_date(start_date)
        end_heb = HebrewDateConverter.get_hebrew_date(end_date)
        
        # Render the print template
        html = render_template('student_report_print.html',
            student_name=f"{student['first_name']} {student['last_name']}",
            student_grade=student.get('current_grade', ''),
            student_birth_date=student.get('birth_date_hebrew', ''),
            student_address=student.get('address', ''),
            student_city=student.get('city', ''),
            student_phone=student.get('home_phone', ''),
            intro_text=intro_text,
            exam_students=exam_students,
            attendance_summary=attendance_summary,
            attendance_weekly=attendance_weekly,
            date_range_start=start_heb,
            date_range_end=end_heb,
            current_date=HebrewDateConverter.get_current_hebrew_date(),
            yeshiva_name='עמשינוב'
        )
        
        return html
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/student-report/print', methods=['POST'])
def student_report_print():
    """פתיחת דוח להדפסה בחלון חדש"""
    try:
        report_data = json.loads(request.form.get('report_data'))
        
        student_id = report_data.get('student_id')
        intro_text = report_data.get('intro_text', '')
        start_date_str = report_data.get('start_date')
        end_date_str = report_data.get('end_date')
        exams = report_data.get('exams', [])
        
        student = db.get_student_by_id(student_id)
        if not student:
            return "תלמיד לא נמצא", 404
        
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y').date()
        
        attendance_summary = db.get_student_attendance_summary(student_id, start_date, end_date)
        attendance_weekly = db.get_student_attendance_weekly(student_id, start_date, end_date)
        
        # Format exam data for horizontal table
        exam_students = [{
            'name': f"{student['first_name']} {student['last_name']}",
            'iyon': exams.get('iyon', ''),
            'bekiut': exams.get('bekiut', ''),
            'gemara_rashi': exams.get('gemara_rashi', ''),
            'chumash': exams.get('chumash', '')
        }]
        
        from services.date_service import HebrewDateConverter
        start_heb = HebrewDateConverter.get_hebrew_date(start_date)
        end_heb = HebrewDateConverter.get_hebrew_date(end_date)
        
        return render_template('student_report_print.html',
            student_name=f"{student['first_name']} {student['last_name']}",
            student_grade=student.get('current_grade', ''),
            student_birth_date=student.get('birth_date_hebrew', ''),
            student_address=student.get('address', ''),
            student_city=student.get('city', ''),
            student_phone=student.get('home_phone', ''),
            intro_text=intro_text,
            exam_students=exam_students,
            attendance_summary=attendance_summary,
            attendance_weekly=attendance_weekly,
            date_range_start=start_heb,
            date_range_end=end_heb,
            current_date=HebrewDateConverter.get_current_hebrew_date(),
            yeshiva_name='עמשינוב'
        )
    except Exception as e:
        return f"שגיאה: {str(e)}", 400

@app.route('/api/student-report/pdf', methods=['POST'])
def api_student_report_pdf():
    """API: יצירת PDF של דוח התלמיד"""
    try:
        data = request.json
        student_id = data.get('student_id')
        intro_text = data.get('intro_text', '')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        exams = data.get('exams', [])
        
        student = db.get_student_by_id(student_id)
        if not student:
            return jsonify({'error': 'תלמיד לא נמצא'}), 404
        
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y').date()
        
        attendance_summary = db.get_student_attendance_summary(student_id, start_date, end_date)
        attendance_weekly = db.get_student_attendance_weekly(student_id, start_date, end_date)
        
        # Format exam data for horizontal table
        exam_students = [{
            'name': f"{student['first_name']} {student['last_name']}",
            'iyon': exams.get('iyon', ''),
            'bekiut': exams.get('bekiut', ''),
            'gemara_rashi': exams.get('gemara_rashi', ''),
            'chumash': exams.get('chumash', '')
        }]
        
        from services.date_service import HebrewDateConverter
        start_heb = HebrewDateConverter.get_hebrew_date(start_date)
        end_heb = HebrewDateConverter.get_hebrew_date(end_date)
        
        html = render_template('student_report_print.html',
            student_name=f"{student['first_name']} {student['last_name']}",
            student_grade=student.get('current_grade', ''),
            student_birth_date=student.get('birth_date_hebrew', ''),
            student_address=student.get('address', ''),
            student_city=student.get('city', ''),
            student_phone=student.get('home_phone', ''),
            intro_text=intro_text,
            exam_students=exam_students,
            attendance_summary=attendance_summary,
            attendance_weekly=attendance_weekly,
            date_range_start=start_heb,
            date_range_end=end_heb,
            current_date=HebrewDateConverter.get_current_hebrew_date(),
            yeshiva_name='עמשינוב'
        )
        
        # Try to create PDF using weasyprint
        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=student_report_{student_id}.pdf'
            return response
        except ImportError:
            # If weasyprint not available, return HTML that can be printed
            response = make_response(html)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            return response
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/students', methods=['POST'])
def api_add_student():
    """API: הוספת תלמיד חדש"""
    try:
        data = request.get_json()
        student_data = {
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'id_number': data.get('id_number', ''),
            'birth_date_hebrew': data.get('birth_date_hebrew', ''),
            'address': data.get('address', ''),
            'city': data.get('city', ''),
            'father_name': data.get('father_name', ''),
            'father_id_number': data.get('father_id_number', ''),
            'mother_name': data.get('mother_name', ''),
            'mother_id_number': data.get('mother_id_number', ''),
            'father_phone': data.get('mobile_phone', ''),  # פלאפון
            'mother_phone': '',
            'home_phone': data.get('home_phone', ''),
            'entry_date_hebrew': '',
            'current_grade': data.get('current_grade', ''),
            'status': 'פעיל',
            'framework_type': 'רגיל',
            'notes': ''
        }

        student_id = db.add_student(student_data)
        return jsonify({'success': True, 'id': student_id}), 201
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/students/all', methods=['DELETE'])
def api_delete_all_students():
    """API: מחיקת כל התלמידים"""
    try:
        db.delete_all_students()
        return jsonify({'success': True, 'message': 'כל התלמידים נמחקו'}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/students/bulk', methods=['POST'])
def api_add_students_bulk():
    """API: הוספה קבוצתית של תלמידים"""
    try:
        students = request.get_json()
        added_count = 0

        for student_data in students:
            try:
                db.add_student(student_data)
                added_count += 1
            except Exception as e:
                print(f"שגיאה בהוספת תלמיד: {e}")
                continue

        return jsonify({'success': True, 'added': added_count}), 201
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/students/import', methods=['POST'])
@login_required
def api_import_students():
    """API: ייבוא תלמידים מקובץ Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'לא נבחר קובץ'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'לא נבחר קובץ'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'פורמט קובץ לא נתמך. יש להעלות קובץ Excel'}), 400
        
        import openpyxl
        from io import BytesIO
        
        # Read Excel file
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Map Hebrew headers to field names
        header_map = {
            'שם פרטי': 'first_name',
            'שם משפחה': 'last_name',
            'תעודת זהות': 'id_number',
            'תאריך לידה': 'birth_date_hebrew',
            'כתובת': 'address',
            'עיר': 'city',
            'שם האב': 'father_name',
            'ת.ז. אב': 'father_id_number',
            'שם האם': 'mother_name',
            'ת.ז. אם': 'mother_id_number',
            'טלפון אב': 'father_phone',
            'טלפון אם': 'mother_phone',
            'טלפון בית': 'home_phone',
            'שיעור': 'current_grade',
            'מסגרת': 'framework_type',
            'הערות': 'notes'
        }
        
        # Create column index map
        col_map = {}
        for i, header in enumerate(headers):
            if header in header_map:
                col_map[header_map[header]] = i
        
        # Check required columns
        if 'first_name' not in col_map or 'last_name' not in col_map:
            return jsonify({
                'success': False, 
                'error': 'חסרות עמודות חובה: שם פרטי, שם משפחה'
            }), 400
        
        # Import students
        added_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                student_data = {}
                for field, col_idx in col_map.items():
                    value = row[col_idx] if col_idx < len(row) else None
                    student_data[field] = str(value).strip() if value else ''
                
                # Skip empty rows
                if not student_data.get('first_name') and not student_data.get('last_name'):
                    continue
                
                # Set defaults
                if not student_data.get('current_grade'):
                    student_data['current_grade'] = "א'"
                if not student_data.get('framework_type'):
                    student_data['framework_type'] = 'ישיבה קטנה'
                
                db.add_student(student_data)
                added_count += 1
                
            except Exception as e:
                errors.append(f"שורה {row_num}: {str(e)}")
        
        result = {
            'success': True,
            'added': added_count,
            'message': f'נוספו {added_count} תלמידים בהצלחה'
        }
        
        if errors:
            result['errors'] = errors[:10]  # Limit errors shown
            result['total_errors'] = len(errors)
        
        return jsonify(result), 201
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/students/template')
@login_required
def api_download_template():
    """API: הורדת תבנית Excel לייבוא תלמידים"""
    import openpyxl
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "תלמידים"
    
    # Add headers
    headers = [
        'שם פרטי', 'שם משפחה', 'תעודת זהות', 'תאריך לידה',
        'כתובת', 'עיר', 'שם האב', 'ת.ז. אב', 'שם האם', 'ת.ז. אם',
        'טלפון אב', 'טלפון אם', 'טלפון בית', 'שיעור', 'מסגרת', 'הערות'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Add example row
    example = ['ישראל', 'ישראלי', '123456789', "א' בתשרי תשפ\"ה",
               'הרצל 1', 'ירושלים', 'אברהם', '111111111', 'שרה', '222222222',
               '050-1234567', '050-7654321', '02-1234567', "א'", 'ישיבה קטנה', '']
    
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='תבנית_ייבוא_תלמידים.xlsx'
    )

# ==================== EXAMS ROUTES ====================

@app.route('/exams')
@login_required
def exams():
    """דף ניהול מבחנים"""
    return render_template('exams/exams.html')

@app.route('/exams/syllabi')
@login_required
def exams_syllabi():
    """דף ניהול הספקים"""
    return render_template('exams/syllabi.html')

@app.route('/exams/create')
@login_required
def exams_create():
    """דף יצירת מבחן"""
    return render_template('exams/create.html')

@app.route('/exams/scan')
@login_required
def exams_scan():
    """דף סריקת מבחנים"""
    return render_template('exams/scan.html')

@app.route('/exams/grades')
@login_required
def exams_grades():
    """דף הזנת ציונים - תצוגת מטריצה"""
    return render_template('exams/grades.html')

# ==================== EXAMS API ====================

@app.route('/api/exams/syllabi', methods=['GET', 'POST'])
def api_syllabi():
    """API לניהול הספקים"""
    if request.method == 'POST':
        data = request.json
        try:
            exam_db.save_syllabus(
                grade=data.get('grade'),
                subject=data.get('subject'),
                content=data.get('content'),
                syllabus_text=data.get('syllabus_text'),
                academic_year=data.get('academic_year'),
                semester=data.get('semester'),
                notes=data.get('notes')
            )
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        grade = request.args.get('grade')
        subject = request.args.get('subject')
        syllabi = exam_db.get_syllabi(grade=grade, subject=subject)

        # המרה לפורמט JSON ידידותי
        result = []
        for row in syllabi:
            result.append({
                'id': row[0],
                'grade': row[1],
                'subject': row[2],
                'content': row[3],
                'syllabus_text': row[4],
                'academic_year': row[5],
                'semester': row[6],
                'notes': row[7]
            })

        return jsonify({'syllabi': result})

@app.route('/api/exams/syllabi/<int:syllabus_id>', methods=['GET', 'PUT', 'DELETE'])
def api_syllabus_detail(syllabus_id):
    """API לניהול סילבוס בודד"""
    if request.method == 'GET':
        # קבלת פרטי סילבוס בודד
        try:
            syllabi = exam_db.get_syllabi()
            syllabus = next((s for s in syllabi if s[0] == syllabus_id), None)

            if not syllabus:
                return jsonify({'error': 'סילבוס לא נמצא'}), 404

            return jsonify({
                'id': syllabus[0],
                'grade': syllabus[1],
                'subject': syllabus[2],
                'content': syllabus[3],
                'syllabus_text': syllabus[4],
                'academic_year': syllabus[5],
                'semester': syllabus[6],
                'notes': syllabus[7]
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    elif request.method == 'PUT':
        # עדכון סילבוס
        data = request.json
        try:
            conn = exam_db.connect()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE subject_syllabi
                SET grade = ?, subject = ?, content = ?, syllabus_text = ?,
                    academic_year = ?, semester = ?, notes = ?
                WHERE id = ?
            ''', (
                data.get('grade'),
                data.get('subject'),
                data.get('content'),
                data.get('syllabus_text'),
                data.get('academic_year'),
                data.get('semester'),
                data.get('notes'),
                syllabus_id
            ))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    elif request.method == 'DELETE':
        # מחיקת סילבוס
        try:
            conn = exam_db.connect()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM subject_syllabi WHERE id = ?', (syllabus_id,))
            conn.commit()
            conn.close()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/api/exams', methods=['GET', 'POST'])
def api_exams():
    """API למבחנים"""
    if request.method == 'POST':
        data = request.json
        try:
            exam_id = exam_db.create_exam(
                exam_data=data.get('exam'),
                questions=data.get('questions', [])
            )
            return jsonify({'success': True, 'exam_id': exam_id})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        grade = request.args.get('grade')
        subject = request.args.get('subject')
        status = request.args.get('status')
        exams = exam_db.get_all_exams(grade=grade, subject=subject, status=status)

        result = []
        for row in exams:
            result.append({
                'id': row[0],
                'subject': row[1],
                'title': row[2],
                'description': row[3],
                'syllabus_text': row[4],
                'grade': row[5],
                'total_points': row[6],
                'created_at': row[7],
                'created_by': row[8],
                'academic_year': row[9],
                'semester': row[10],
                'status': row[11]
            })

        return jsonify(result)

@app.route('/api/exams/<int:exam_id>')
def api_exam_detail(exam_id):
    """פרטי מבחן"""
    exam = exam_db.get_exam(exam_id)
    questions = exam_db.get_exam_questions(exam_id)

    if not exam:
        return jsonify({'error': 'מבחן לא נמצא'}), 404

    return jsonify({
        'id': exam[0],
        'subject': exam[1],
        'title': exam[2],
        'description': exam[3],
        'syllabus_text': exam[4],
        'grade': exam[5],
        'total_points': exam[6],
        'created_at': exam[7],
        'created_by': exam[8],
        'academic_year': exam[9],
        'semester': exam[10],
        'status': exam[11],
        'questions': [{
            'id': q[0],
            'exam_id': q[1],
            'question_number': q[2],
            'question_text': q[3],
            'points': q[4],
            'question_type': q[5],
            'correct_answer': q[6]
        } for q in questions]
    })

@app.route('/api/exams/<int:exam_id>/assign', methods=['POST'])
def api_assign_exam(exam_id):
    """הקצאת מבחן לתלמידים"""
    data = request.json
    try:
        assigned_ids = exam_db.assign_exam_to_students(
            exam_id=exam_id,
            student_ids=data.get('student_ids', []),
            scheduled_date=data.get('scheduled_date'),
            version_code=data.get('version', 'A')
        )
        return jsonify({'success': True, 'count': len(assigned_ids)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exams/grades', methods=['POST'])
def api_save_grade():
    """שמירת ציון"""
    data = request.json
    try:
        grade_id = exam_db.save_exam_grade(
            student_exam_id=data.get('student_exam_id'),
            total_score=data.get('total_score'),
            graded_by=data.get('graded_by', 'מערכת'),
            grading_method=data.get('grading_method', 'manual'),
            ocr_confidence=data.get('ocr_confidence'),
            notes=data.get('notes')
        )
        return jsonify({'success': True, 'grade_id': grade_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/exams/grades/matrix')
def api_grades_matrix():
    """מטריצת ציונים - שורות=תלמידים, עמודות=מבחנים"""
    try:
        grade = request.args.get('grade')
        subject = request.args.get('subject')
        matrix = exam_db.get_grades_matrix(grade=grade, subject=subject)
        return jsonify(matrix)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/grades/direct', methods=['POST'])
def api_save_grade_direct():
    """שמירת ציון ישירה לפי student_id ו-exam_id"""
    data = request.json
    try:
        result = exam_db.save_grade_direct(
            student_id=data.get('student_id'),
            exam_id=data.get('exam_id'),
            score=data.get('score'),
            graded_by=data.get('graded_by', 'מערכת')
        )
        if result:
            return jsonify({'success': True, **result})
        else:
            return jsonify({'success': False, 'error': 'מבחן לא נמצא'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/students/<int:student_id>/grades')
def api_student_grades(student_id):
    """ציוני תלמיד"""
    grades = exam_db.get_student_grades(student_id)

    result = []
    for row in grades:
        result.append({
            'title': row[0],
            'subject': row[1],
            'scheduled_date': row[2],
            'actual_date': row[3],
            'total_score': row[4],
            'grade_percent': row[5],
            'status': row[6],
            'total_points': row[7]
        })

    return jsonify(result)

@app.route('/api/exams/<int:exam_id>/statistics')
def api_exam_statistics(exam_id):
    """סטטיסטיקות מבחן"""
    stats = exam_db.get_exam_statistics(exam_id)
    distribution = exam_db.get_grade_distribution(exam_id)

    return jsonify({
        'total_students': stats[0],
        'graded_count': stats[1],
        'avg_percent': round(stats[2], 2) if stats[2] else 0,
        'min_percent': stats[3],
        'max_percent': stats[4],
        'distribution': [{'range': d[0], 'count': d[1]} for d in distribution]
    })

@app.route('/api/exams/<int:exam_id>', methods=['PUT'])
def api_update_exam(exam_id):
    """עדכון מבחן"""
    data = request.json
    try:
        conn = exam_db.connect()
        cursor = conn.cursor()

        exam_data = data.get('exam', {})

        # עדכון פרטי המבחן
        cursor.execute('''
            UPDATE exams
            SET subject = ?, title = ?, description = ?, syllabus_text = ?,
                grade = ?, total_points = ?, academic_year = ?, semester = ?, status = ?
            WHERE id = ?
        ''', (
            exam_data.get('subject'),
            exam_data.get('title'),
            exam_data.get('description'),
            exam_data.get('syllabus_text'),
            exam_data.get('grade'),
            exam_data.get('total_points'),
            exam_data.get('academic_year'),
            exam_data.get('semester'),
            exam_data.get('status'),
            exam_id
        ))

        # מחיקת שאלות קיימות והוספת חדשות
        cursor.execute('DELETE FROM exam_questions WHERE exam_id = ?', (exam_id,))

        questions = data.get('questions', [])
        for q in questions:
            cursor.execute('''
                INSERT INTO exam_questions (exam_id, question_number, question_text, points, question_type, correct_answer)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (exam_id, q['question_number'], q['question_text'], q['points'], q['question_type'], q.get('correct_answer')))

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'exam_id': exam_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/<int:exam_id>', methods=['DELETE'])
def api_delete_exam(exam_id):
    """מחיקת מבחן"""
    try:
        conn = exam_db.connect()
        cursor = conn.cursor()

        # מחיקת כל הנתונים הקשורים למבחן
        cursor.execute('DELETE FROM question_grades WHERE student_exam_id IN (SELECT id FROM student_exams WHERE exam_id = ?)', (exam_id,))
        cursor.execute('DELETE FROM exam_grades WHERE student_exam_id IN (SELECT id FROM student_exams WHERE exam_id = ?)', (exam_id,))
        cursor.execute('DELETE FROM student_exams WHERE exam_id = ?', (exam_id,))
        cursor.execute('DELETE FROM exam_questions WHERE exam_id = ?', (exam_id,))
        cursor.execute('DELETE FROM exam_versions WHERE exam_id = ?', (exam_id,))
        cursor.execute('DELETE FROM exams WHERE id = ?', (exam_id,))

        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/stats')
def api_exams_stats():
    """סטטיסטיקות כלליות"""
    try:
        conn = exam_db.connect()
        cursor = conn.cursor()

        # מבחנים מתוכננים
        cursor.execute("SELECT COUNT(*) FROM exams WHERE status = 'scheduled'")
        scheduled = cursor.fetchone()[0]

        # מבחנים שהושלמו החודש
        cursor.execute("""
            SELECT COUNT(*) FROM exams
            WHERE status = 'completed'
            AND strftime('%Y-%m', created_at) = strftime('%Y-%m', 'now')
        """)
        completed = cursor.fetchone()[0]

        # ממוצע ציונים
        cursor.execute("SELECT AVG(total_score / total_points * 100) FROM exam_grades eg JOIN student_exams se ON eg.student_exam_id = se.id JOIN exams e ON se.exam_id = e.id")
        avg_grade = cursor.fetchone()[0]

        # מספר תלמידים שבוחנו
        cursor.execute("SELECT COUNT(DISTINCT se.student_id) FROM student_exams se JOIN exam_grades eg ON se.id = eg.student_exam_id")
        students_tested = cursor.fetchone()[0]

        conn.close()

        return jsonify({
            'scheduled': scheduled or 0,
            'completed': completed or 0,
            'avg_grade': round(avg_grade, 1) if avg_grade else 0,
            'students_tested': students_tested or 0
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/<int:exam_id>/pdf')
def api_exam_pdf(exam_id):
    """הורדת PDF של מבחן"""
    student_id = request.args.get('student_id')

    if not student_id:
        return jsonify({'error': 'student_id is required'}), 400

    try:
        from services.pdf_generator import generate_exam_pdf_for_student
        from io import BytesIO

        # יצירת ה-PDF
        pdf_bytes = generate_exam_pdf_for_student(exam_id, int(student_id), exam_db)

        # החזרת הקובץ
        return send_file(
            BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'exam_{exam_id}_student_{student_id}.pdf'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/<int:exam_id>/pdf/batch', methods=['POST'])
def api_exam_pdf_batch(exam_id):
    """יצירת PDFs לכל התלמידים"""
    data = request.json
    student_ids = data.get('student_ids', [])

    if not student_ids:
        return jsonify({'error': 'student_ids are required'}), 400

    try:
        from services.pdf_generator import generate_batch_pdfs
        import tempfile
        import zipfile
        from io import BytesIO
        import os

        # יצירת תיקייה זמנית
        with tempfile.TemporaryDirectory() as temp_dir:
            # יצירת PDFs
            pdf_files = generate_batch_pdfs(exam_id, student_ids, temp_dir, exam_db)

            # יצירת ZIP
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for pdf_path in pdf_files:
                    zip_file.write(pdf_path, os.path.basename(pdf_path))

            zip_buffer.seek(0)

            # החזרת הקובץ
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name=f'exam_{exam_id}_all_students.zip'
            )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== SCANNING API ====================

@app.route('/api/exams/scan/upload', methods=['POST'])
def api_scan_upload():
    """העלאת קבצים לסריקה"""
    print("\n" + "="*60)
    print("DEBUG: API scan upload endpoint called")

    if 'files' not in request.files:
        print("DEBUG: No files in request")
        return jsonify({'error': 'No files provided'}), 400

    files = request.files.getlist('files')
    print(f"DEBUG: Got {len(files)} files")

    if not files:
        print("DEBUG: Files list is empty")
        return jsonify({'error': 'No files provided'}), 400

    try:
        from services.ocr_service import ExamOCRService

        print("DEBUG: Creating OCR service...")
        ocr = ExamOCRService(tesseract_path=app.config.get('TESSERACT_PATH'))
        all_results = []

        for file in files:
            print(f"\nDEBUG: Processing file: {file.filename}")
            if file.filename == '':
                print("DEBUG: Empty filename, skipping")
                continue

            # שמירה זמנית
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            print(f"DEBUG: Saving file to: {filepath}")
            file.save(filepath)
            print(f"DEBUG: File saved successfully, size: {os.path.getsize(filepath)} bytes")

            try:
                # עיבוד לפי סוג הקובץ
                if filename.lower().endswith('.pdf'):
                    print(f"DEBUG: Processing as PDF...")
                    results = ocr.process_pdf(filepath)
                    print(f"DEBUG: PDF processing returned {len(results)} results")
                else:
                    # תמונה בודדת
                    from PIL import Image
                    img = Image.open(filepath)
                    result = ocr.process_single_page(img)
                    result['file_name'] = filename
                    results = [result]

                # הוספה לתוצאות
                for r in results:
                    r['file_name'] = filename

                    # אם יש qr_data עם student_id, שלוף את שם התלמיד מבסיס הנתונים
                    if r.get('qr_data') and 'student_id' in r['qr_data']:
                        student_id = r['qr_data']['student_id']
                        exam_id = r['qr_data'].get('exam_id')

                        # שליפת שם התלמיד (שם משפחה + שם פרטי)
                        student = db.get_student(student_id)
                        if student:
                            r['qr_data']['student_name'] = f"{student['last_name']} {student['first_name']}"
                        else:
                            r['qr_data']['student_name'] = f"תלמיד #{student_id} (לא נמצא)"

                        # שליפת שם המבחן
                        if exam_id:
                            exam = exam_db.get_exam(exam_id)
                            if exam:
                                r['qr_data']['exam_title'] = exam['title']

                    all_results.append(r)

            finally:
                # מחיקת הקובץ הזמני
                if os.path.exists(filepath):
                    os.remove(filepath)

        return jsonify({'results': all_results})

    except Exception as e:
        print(f"ERROR in scan upload: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/exams/scan/save', methods=['POST'])
def api_scan_save():
    """שמירת ציונים מסריקה"""
    data = request.json
    results = data.get('results', [])

    if not results:
        return jsonify({'error': 'No results to save'}), 400

    try:
        from services.ocr_service import save_grades_to_database

        stats = save_grades_to_database(results, exam_db)

        return jsonify(stats)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== DEVELOPER FEEDBACK ====================

@app.route('/feedback')
@login_required
def feedback_page():
    """דף משוב למתכנת"""
    return render_template('feedback.html')

@app.route('/api/feedback', methods=['GET', 'POST'])
@login_required
def api_feedback():
    """API למשוב"""
    if request.method == 'POST':
        data = request.json
        try:
            feedback_id = db.add_feedback(
                category=data.get('category', 'כללי'),
                title=data.get('title', ''),
                description=data.get('description', ''),
                priority=data.get('priority', 'רגיל')
            )
            return jsonify({'success': True, 'id': feedback_id})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        status = request.args.get('status')
        feedback_list = db.get_all_feedback(status=status)
        return jsonify(feedback_list)

@app.route('/api/feedback/<int:feedback_id>', methods=['PUT', 'DELETE'])
@login_required
def api_feedback_item(feedback_id):
    """API לעדכון/מחיקת משוב"""
    if request.method == 'DELETE':
        try:
            db.delete_feedback(feedback_id)
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:  # PUT
        data = request.json
        try:
            db.update_feedback_status(
                feedback_id=feedback_id,
                status=data.get('status'),
                notes=data.get('notes')
            )
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

# ==================== FAVICON ====================

@app.route('/favicon.ico')
def favicon():
    """Favicon endpoint - returns 204 No Content to avoid errors"""
    return '', 204

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    return render_template('dashboard.html'), 404

@app.errorhandler(500)
def server_error(error):
    return jsonify({'error': 'שגיאה בשרת'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
